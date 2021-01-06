# importing libraries for Excel files
import openpyxl as xl
from openpyxl import Workbook
from copy import copy

# importing libraries to extract url data
import requests
from bs4 import BeautifulSoup

# importing libraries for automation
from datetime import datetime, timedelta
from threading import Timer

# opening the source excel file
source_file = "C:\\Users\\Hp\\Desktop\\Gold and Silver Performance Tracker\\Analysis.xlsx"
wb1 = xl.load_workbook(source_file)
ws1 = wb1.worksheets[0]

# creating the new excel file
new_file = Workbook()
ws2 = new_file.active

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

# copying the cell values from source
# excel file to destination excel file
for i in range(1, mr + 1):
    for j in range(1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row=i, column=j)

        # writing the read value to destination excel file
        p = ws2.cell(row=i, column=j)
        p.value = c.value

        # Copying and pasting the cell style from source file to the new file
        if c.has_style:
            p.font = copy(c.font)
            p.border = copy(c.border)
            p.fill = copy(c.fill)
            p.number_format = copy(c.number_format)
            p.protection = copy(c.protection)
            p.alignment = copy(c.alignment)

# adjusting the column sizes
'''for column_cells in ws2.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws2.column_dimensions[column_cells[0].column_letter].width = length'''

# Set page orientation to landscape
ws2.page_setup.orientation = ws2.ORIENTATION_LANDSCAPE


# importing data from the website (silver/gold rates)

# accessing website and its content
url_silver = 'http://www.livepriceofgold.com/silver-price/malaysia.html'
url_gold = 'http://www.livepriceofgold.com/malaysia-gold-price.html'
response_silver = requests.get(url_silver)
response_gold = requests.get(url_gold)

# parse the html content (text) into a list
soup_silver = BeautifulSoup(response_silver.text, "html.parser")
soup_gold = BeautifulSoup(response_gold.text, "html.parser")

# find the location of required value from list
# td is the general html tag of where the value is
# 10 is the location of the value from the list
scraped_silver = soup_silver.findAll('td')[10]
scraped_gold = soup_gold.findAll('td')[10]

# Get the values and convert to int type.
scraped_silver = str(scraped_silver)
strvalue_silver = scraped_silver[scraped_silver.find('>') + 1:scraped_silver.find('</td>')]
intvalue_silver = float(strvalue_silver)
print(f'Spot Silver Price in MYR: {intvalue_silver}')

scraped_gold = str(scraped_gold)
strvalue_gold = scraped_gold[scraped_gold.find('>') + 1:scraped_gold.find('</td>')]
intvalue_gold = float(strvalue_gold)
print(f'Gold Rate per Gram in MYR: {intvalue_gold}')

# Replace the cell values (Market price current and No. of months elapsed)
a = 6
for row in ws2['K6:K10']:
    for cell in row:
        cell.value = intvalue_silver * ws2.cell(row=a, column=1).value
    a += 1

b = 19
for row in ws2['K19:K21']:
    for cell in row:
        cell.value = intvalue_gold * ws2.cell(row=b, column=1).value
    b += 1

present_date = datetime.today()
c = 6
for row in ws2['C6:C10']:
    for cell in row:
        prev_date_s = ws2.cell(row=c, column=2).value
        delta_s = (present_date.year - prev_date_s.year) * 12 + (present_date.month - prev_date_s.month)
        cell.value = delta_s
    c += 1

d = 19
for row in ws2['C19:C21']:
    for cell in row:
        prev_date_g = ws2.cell(row=d, column=2).value
        delta_g = (present_date.year - prev_date_g.year) * 12 + (present_date.month - prev_date_g.month)
        cell.value = delta_g
    d += 1

# function to save the excel file into a specific directory


def xlfile():
    now = datetime.now()
    str_date_today = now.strftime("%Y_%m_%d %H_%M_%S")
    file_path = "C:\\Users\\Hp\Desktop\\Gold and Silver Performance Tracker\\"
    file_title = "G & S " + str_date_today
    file_type = '.xlsx'
    full_filename = file_path + file_title + file_type
    new_file.save(full_filename)


# set the time for automated excel file saving (only works when program is running)
'''today = datetime.today()
checkpoint = today.replace(day=today.day, hour=12, minute=00, second=0, microsecond=0) + timedelta(days=1)
delta_t = checkpoint-today

secs = delta_t.total_seconds()
t = Timer(secs, xlfile)
t.start()'''


# save the file when wanted
xlfile()
